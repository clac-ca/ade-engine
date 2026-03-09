from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook

from ade_engine.infrastructure.io.workbook import _convert_xlrd_book_to_openpyxl, load_source_workbook, resolve_sheet_names
from ade_engine.infrastructure.settings import Settings
from ade_engine.models.errors import InputError


@dataclass
class _FakeCell:
    ctype: int
    value: object


class _FakeSheet:
    def __init__(
        self,
        name: str,
        rows: list[list[_FakeCell]],
        *,
        merged_cells: list[tuple[int, int, int, int]] | None = None,
        visibility: int | None = None,
        sheet_visible: int = 0,
        sheet_selected: int = 0,
    ) -> None:
        self.name = name
        self._rows = rows
        self.nrows = len(rows)
        self.ncols = max((len(row) for row in rows), default=0)
        self.merged_cells = merged_cells or []
        if visibility is not None:
            self.visibility = visibility
        self.sheet_visible = sheet_visible
        self.sheet_selected = sheet_selected

    def cell(self, row_index: int, col_index: int) -> _FakeCell:
        if row_index >= self.nrows or col_index >= len(self._rows[row_index]):
            return _FakeCell(0, "")
        return self._rows[row_index][col_index]

    def row_len(self, row_index: int) -> int:
        return len(self._rows[row_index])

    def row_types(self, row_index: int, start_colx: int = 0, end_colx: int | None = None) -> list[int]:
        row = self._rows[row_index]
        end = len(row) if end_colx is None else end_colx
        return [cell.ctype for cell in row[start_colx:end]]

    def row_values(self, row_index: int, start_colx: int = 0, end_colx: int | None = None) -> list[object]:
        row = self._rows[row_index]
        end = len(row) if end_colx is None else end_colx
        return [cell.value for cell in row[start_colx:end]]


class _FakeBook:
    def __init__(self, sheets: list[_FakeSheet], *, datemode: int = 0, sheet_visibility: list[int] | None = None) -> None:
        self._sheets = sheets
        self.nsheets = len(sheets)
        self.datemode = datemode
        self.sheet_visibility = sheet_visibility or [0] * len(sheets)
        self._sheet_visibility = self.sheet_visibility

    def sheet_by_index(self, index: int) -> _FakeSheet:
        return self._sheets[index]


def _write_xls(path: Path) -> None:
    xlwt = pytest.importorskip("xlwt")

    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("Visible")
    sheet.write_merge(0, 0, 0, 1, "Header")
    sheet.write(1, 0, "USER@Example.com")
    sheet.write(1, 1, "Alice")
    workbook.save(str(path))


def test_settings_default_supports_xls():
    assert ".xls" in Settings().supported_file_extensions


def test_convert_xlrd_book_to_openpyxl_preserves_visibility_merges_and_cell_types():
    xlrd = pytest.importorskip("xlrd")

    visible_sheet = _FakeSheet(
        "Visible",
        rows=[
            [_FakeCell(xlrd.XL_CELL_TEXT, "Header"), _FakeCell(xlrd.XL_CELL_EMPTY, "")],
            [
                _FakeCell(xlrd.XL_CELL_NUMBER, 4.0),
                _FakeCell(xlrd.XL_CELL_NUMBER, 4.5),
            ],
            [
                _FakeCell(xlrd.XL_CELL_BOOLEAN, 1),
                _FakeCell(xlrd.XL_CELL_DATE, 0.0),
            ],
            [
                _FakeCell(xlrd.XL_CELL_ERROR, 0x0F),
                _FakeCell(xlrd.XL_CELL_BLANK, ""),
            ],
        ],
        merged_cells=[(0, 1, 0, 2)],
        sheet_visible=1,
    )
    hidden_sheet = _FakeSheet("Hidden", rows=[[_FakeCell(xlrd.XL_CELL_TEXT, "secret")]], visibility=1)
    book = _FakeBook([visible_sheet, hidden_sheet], sheet_visibility=[0, 1])

    workbook = _convert_xlrd_book_to_openpyxl(book)

    assert workbook.sheetnames == ["Visible", "Hidden"]
    assert workbook["Visible"].sheet_state == "visible"
    assert workbook["Hidden"].sheet_state == "hidden"
    assert workbook.active.title == "Visible"
    assert "A1:B1" in {str(cell_range) for cell_range in workbook["Visible"].merged_cells.ranges}
    assert workbook["Visible"]["A2"].value == 4
    assert workbook["Visible"]["B2"].value == 4.5
    assert workbook["Visible"]["A3"].value is True
    assert workbook["Visible"]["B3"].value == datetime(1899, 12, 31, 0, 0)
    assert workbook["Visible"]["A4"].value == "#VALUE!"
    assert workbook["Visible"]["B4"].value is None


def test_load_source_workbook_reads_csv_into_workbook(tmp_path: Path):
    source = tmp_path / "input.csv"
    source.write_text("Email,Name\nUSER@Example.com,Alice\n", encoding="utf-8")

    workbook = load_source_workbook(source)

    assert workbook.sheetnames == ["input"]
    assert workbook.active["A2"].value == "USER@Example.com"
    workbook.close()


def test_load_source_workbook_reads_xls_into_workbook(tmp_path: Path):
    source = tmp_path / "input.xls"
    _write_xls(source)

    workbook = load_source_workbook(source)

    assert workbook.sheetnames == ["Visible"]
    assert workbook.active["A1"].value == "Header"
    assert workbook.active["A2"].value == "USER@Example.com"
    workbook.close()


def test_convert_xlrd_book_to_openpyxl_uses_row_arrays_when_cell_access_breaks():
    xlrd = pytest.importorskip("xlrd")

    class _BrokenCellSheet(_FakeSheet):
        def cell(self, row_index: int, col_index: int) -> _FakeCell:
            raise IndexError("array index out of range")

    book = _FakeBook(
        [
            _BrokenCellSheet(
                "Visible",
                rows=[
                    [_FakeCell(xlrd.XL_CELL_TEXT, "Header")],
                    [_FakeCell(xlrd.XL_CELL_NUMBER, 4.0)],
                ],
            )
        ]
    )

    workbook = _convert_xlrd_book_to_openpyxl(book)

    assert workbook.active["A1"].value == "Header"
    assert workbook.active["A2"].value == 4


def test_convert_xlrd_book_to_openpyxl_uses_sheet_visible_for_active_sheet():
    xlrd = pytest.importorskip("xlrd")

    book = _FakeBook(
        [
            _FakeSheet("First", rows=[[_FakeCell(xlrd.XL_CELL_TEXT, "a")]]),
            _FakeSheet("Second", rows=[[_FakeCell(xlrd.XL_CELL_TEXT, "b")]], sheet_visible=1),
            _FakeSheet("Hidden", rows=[[_FakeCell(xlrd.XL_CELL_TEXT, "c")]], visibility=1, sheet_visible=1),
        ]
    )

    workbook = _convert_xlrd_book_to_openpyxl(book)

    assert workbook.active.title == "Second"
    assert resolve_sheet_names(workbook, requested=None, active_only=True) == ["Second"]


def test_convert_xlrd_book_to_openpyxl_falls_back_to_first_visible_sheet():
    xlrd = pytest.importorskip("xlrd")

    book = _FakeBook(
        [
            _FakeSheet("First", rows=[[_FakeCell(xlrd.XL_CELL_TEXT, "a")]]),
            _FakeSheet("Hidden", rows=[[_FakeCell(xlrd.XL_CELL_TEXT, "b")]], visibility=1),
            _FakeSheet("Third", rows=[[_FakeCell(xlrd.XL_CELL_TEXT, "c")]], sheet_selected=1),
        ]
    )

    workbook = _convert_xlrd_book_to_openpyxl(book)

    assert workbook.active.title == "Third"

    book = _FakeBook(
        [
            _FakeSheet("First", rows=[[_FakeCell(xlrd.XL_CELL_TEXT, "a")]]),
            _FakeSheet("Hidden", rows=[[_FakeCell(xlrd.XL_CELL_TEXT, "b")]], visibility=1),
            _FakeSheet("Third", rows=[[_FakeCell(xlrd.XL_CELL_TEXT, "c")]]),
        ]
    )
    workbook = _convert_xlrd_book_to_openpyxl(book)

    assert workbook.active.title == "First"


def test_resolve_sheet_names_for_xls_uses_original_names_when_openpyxl_renames_duplicates():
    xlrd = pytest.importorskip("xlrd")

    book = _FakeBook(
        [
            _FakeSheet("Sheet", rows=[[_FakeCell(xlrd.XL_CELL_TEXT, "a")]]),
            _FakeSheet("sheet", rows=[[_FakeCell(xlrd.XL_CELL_TEXT, "b")]]),
        ]
    )

    workbook = _convert_xlrd_book_to_openpyxl(book)

    assert workbook.sheetnames == ["Sheet", "sheet1"]
    assert resolve_sheet_names(workbook, requested=["sheet"]) == ["sheet1"]


def test_resolve_sheet_names_for_xls_excludes_hidden_requested_sheets():
    xlrd = pytest.importorskip("xlrd")

    book = _FakeBook(
        [
            _FakeSheet("Visible", rows=[[_FakeCell(xlrd.XL_CELL_TEXT, "a")]]),
            _FakeSheet("Hidden", rows=[[_FakeCell(xlrd.XL_CELL_TEXT, "b")]], visibility=1),
        ]
    )

    workbook = _convert_xlrd_book_to_openpyxl(book)

    assert resolve_sheet_names(workbook, requested=None) == ["Visible"]
    with pytest.raises(InputError, match="Worksheet\\(s\\) not found: Hidden"):
        resolve_sheet_names(workbook, requested=["Hidden"])


def test_resolve_sheet_names_rejects_active_only_when_no_visible_sheets():
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Hidden"
    sheet.sheet_state = "hidden"

    with pytest.raises(InputError, match="No visible worksheets available"):
        resolve_sheet_names(workbook, requested=None, active_only=True)

