from __future__ import annotations

from openpyxl import Workbook
import polars as pl
import pytest

from ade_engine.application.pipeline.pipeline import Pipeline, _build_merged_table_result, _merge_tables_in_sheet
from ade_engine.extensions.registry import Registry
from ade_engine.infrastructure.observability.logger import NullLogger
from ade_engine.infrastructure.settings import Settings
from ade_engine.models.errors import PipelineError
from ade_engine.models.extension_contexts import FieldDef, RowKind
from ade_engine.models.table import MappedColumn, SourceColumn, TableRegion, TableResult


def test_process_sheet_renders_multiple_tables_with_blank_row():
    registry = Registry()
    logger = NullLogger()

    def detector(*, row_index, **_):
        if row_index in (1, 4):
            return {RowKind.HEADER.value: 1.0}
        if row_index in (2, 3, 5):
            return {RowKind.DATA.value: 1.0}
        return {}

    registry.register_row_detector(detector, row_kind=RowKind.UNKNOWN.value, priority=0)
    registry.finalize()

    pipeline = Pipeline(registry=registry, settings=Settings(), logger=logger)

    source_wb = Workbook()
    source_ws = source_wb.active
    source_ws.title = "Sheet1"
    source_ws.append(["A", "B"])
    source_ws.append([1, 2])
    source_ws.append([3, 4])
    source_ws.append(["C", "D"])
    source_ws.append([5, 6])

    output_wb = Workbook()
    output_wb.remove(output_wb.active)
    output_ws = output_wb.create_sheet(title="Sheet1")

    tables = pipeline.process_sheet(
        sheet=source_ws,
        output_sheet=output_ws,
        state={},
        metadata={"input_file": "input.xlsx", "sheet_index": 0},
        input_file_name="input.xlsx",
    )

    assert [t.table_index for t in tables] == [0, 1]

    emitted = list(output_ws.iter_rows(min_row=1, max_row=6, max_col=2, values_only=True))
    assert emitted == [
        ("A", "B"),
        (1, 2),
        (3, 4),
        (None, None),  # blank separator row
        ("C", "D"),
        (5, 6),
    ]


def test_process_sheet_sorts_tables_by_mapping_ratio():
    registry = Registry()
    logger = NullLogger()

    registry.register_field(FieldDef(name="email"))
    registry.register_field(FieldDef(name="name"))

    def detector(*, row_index, **_):
        if row_index in (1, 4):
            return {RowKind.HEADER.value: 1.0}
        if row_index in (2, 5):
            return {RowKind.DATA.value: 1.0}
        return {}

    def detect_email(*, column_header_original, **_):
        return {"email": 1.0} if column_header_original.strip().lower() == "email" else {}

    def detect_name(*, column_header_original, **_):
        return {"name": 1.0} if column_header_original.strip().lower() == "name" else {}

    registry.register_row_detector(detector, row_kind=RowKind.UNKNOWN.value, priority=0)
    registry.register_column_detector(detect_email, field="email", priority=0)
    registry.register_column_detector(detect_name, field="name", priority=0)
    registry.finalize()

    pipeline = Pipeline(
        registry=registry,
        settings=Settings(sort_tables_by_mapping_ratio=True),
        logger=logger,
    )

    source_wb = Workbook()
    source_ws = source_wb.active
    source_ws.title = "Sheet1"
    source_ws.append(["Email", "Notes"])
    source_ws.append(["a@example.com", "first"])
    source_ws.append([])
    source_ws.append(["Email", "Name"])
    source_ws.append(["b@example.com", "Bob"])

    output_wb = Workbook()
    output_wb.remove(output_wb.active)
    output_ws = output_wb.create_sheet(title="Sheet1")

    tables = pipeline.process_sheet(
        sheet=source_ws,
        output_sheet=output_ws,
        state={},
        metadata={"input_file": "input.xlsx", "sheet_index": 0},
        input_file_name="input.xlsx",
    )

    assert [t.table_index for t in tables] == [1, 0]

    emitted = list(output_ws.iter_rows(min_row=1, max_row=5, max_col=2, values_only=True))
    assert emitted == [
        ("email", "name"),
        ("b@example.com", "Bob"),
        (None, None),  # blank separator row
        ("email", "Notes"),
        ("a@example.com", "first"),
    ]


def test_process_sheet_handles_mixed_numeric_types():
    registry = Registry()
    logger = NullLogger()

    def detector(*, row_index, **_):
        if row_index == 1:
            return {RowKind.HEADER.value: 1.0}
        return {RowKind.DATA.value: 1.0}

    registry.register_row_detector(detector, row_kind=RowKind.UNKNOWN.value, priority=0)
    registry.finalize()

    pipeline = Pipeline(registry=registry, settings=Settings(), logger=logger)

    source_wb = Workbook()
    source_ws = source_wb.active
    source_ws.title = "Sheet1"
    source_ws.append(["Amount"])
    source_ws.append([170])
    source_ws.append([169.75])

    output_wb = Workbook()
    output_wb.remove(output_wb.active)
    output_ws = output_wb.create_sheet(title="Sheet1")

    pipeline.process_sheet(
        sheet=source_ws,
        output_sheet=output_ws,
        state={},
        metadata={"input_file": "input.xlsx", "sheet_index": 0},
        input_file_name="input.xlsx",
    )

    emitted = list(output_ws.iter_rows(min_row=1, max_row=3, max_col=1, values_only=True))
    assert emitted == [("Amount",), (170.0,), (169.75,)]


def test_on_table_written_receives_written_table_after_output_policies():
    """The on_table_written hook should see exactly what was written to the output sheet."""

    registry = Registry()
    logger = NullLogger()

    registry.register_field(FieldDef(name="email"))
    registry.register_field(FieldDef(name="name"))

    def row_detector(*, row_index, **_):
        if row_index == 1:
            return {RowKind.HEADER.value: 1.0}
        return {RowKind.DATA.value: 1.0}

    def detect_email(*, column_header_original, **_):
        return {"email": 1.0} if column_header_original.strip().lower() == "email" else {}

    def detect_name(*, column_header_original, **_):
        return {"name": 1.0} if column_header_original.strip().lower() == "name" else {}

    def capture_written(*, write_table, output_sheet, table_result, state, **_):
        state["written_columns"] = list(write_table.columns)
        state["written_sheet_title"] = getattr(output_sheet, "title", getattr(output_sheet, "name", ""))
        state["output_region_a1"] = table_result.output_region.a1 if table_result.output_region else None
        state["mapped_headers"] = [col.header for col in table_result.mapped_columns]

    registry.register_row_detector(row_detector, row_kind=RowKind.UNKNOWN.value, priority=0)
    registry.register_column_detector(detect_email, field="email", priority=0)
    registry.register_column_detector(detect_name, field="name", priority=0)
    registry.register_hook(capture_written, hook="on_table_written", priority=0)
    registry.finalize()

    pipeline = Pipeline(
        registry=registry,
        settings=Settings(
            remove_unmapped_columns=True,
            write_diagnostics_columns=False,
        ),
        logger=logger,
    )

    source_wb = Workbook()
    source_ws = source_wb.active
    source_ws.title = "Sheet1"
    source_ws.append(["Email", "Name", "Notes"])
    source_ws.append(["a@example.com", "Alice", "keep?"])
    source_ws.append(["b@example.com", "Bob", "keep?"])

    output_wb = Workbook()
    output_wb.remove(output_wb.active)
    output_ws = output_wb.create_sheet(title="Sheet1")

    state: dict = {}
    pipeline.process_sheet(
        sheet=source_ws,
        output_sheet=output_ws,
        state=state,
        metadata={"input_file": "input.xlsx", "sheet_index": 0},
        input_file_name="input.xlsx",
    )

    assert state["written_sheet_title"] == "Sheet1"
    assert state["written_columns"] == ["email", "name"]
    assert state["output_region_a1"] == "A1:B3"
    assert state["mapped_headers"] == ["Email", "Name"]

    headers = list(output_ws.iter_rows(min_row=1, max_row=1, max_col=2, values_only=True))[0]
    assert headers == ("email", "name")


def test_merge_tables_in_sheet_merges_metadata_from_all_tables():
    registry = Registry()
    logger = NullLogger()

    registry.register_field(FieldDef(name="email"))
    registry.register_field(FieldDef(name="name"))

    def detector(*, row_index, **_):
        if row_index in (1, 4):
            return {RowKind.HEADER.value: 1.0}
        if row_index in (2, 5):
            return {RowKind.DATA.value: 1.0}
        return {}

    def detect_email(*, column_header_original, **_):
        return {"email": 1.0} if column_header_original.strip().lower() == "email" else {}

    def detect_name(*, column_header_original, **_):
        return {"name": 1.0} if column_header_original.strip().lower() == "name" else {}

    registry.register_row_detector(detector, row_kind=RowKind.UNKNOWN.value, priority=0)
    registry.register_column_detector(detect_email, field="email", priority=0)
    registry.register_column_detector(detect_name, field="name", priority=0)
    registry.finalize()

    pipeline = Pipeline(
        registry=registry,
        settings=Settings(merge_tables_in_sheet=True, write_diagnostics_columns=False),
        logger=logger,
    )

    source_wb = Workbook()
    source_ws = source_wb.active
    source_ws.title = "Sheet1"
    source_ws.append(["Email"])
    source_ws.append(["a@example.com"])
    source_ws.append([])
    source_ws.append(["Name"])
    source_ws.append(["Alice"])

    output_wb = Workbook()
    output_wb.remove(output_wb.active)
    output_ws = output_wb.create_sheet(title="Sheet1")

    tables = pipeline.process_sheet(
        sheet=source_ws,
        output_sheet=output_ws,
        state={},
        metadata={"input_file": "input.xlsx", "sheet_index": 0},
        input_file_name="input.xlsx",
    )

    assert len(tables) == 1
    table = tables[0]
    assert "email" in table.table.columns
    assert "name" in table.table.columns
    assert table.source_region.a1 == "A1:A5"
    assert {column.field_name for column in table.mapped_columns} == {"email", "name"}
    mapped_indices = {column.field_name: column.source_index for column in table.mapped_columns}
    assert table.table.columns[mapped_indices["email"]] == "email"
    assert table.table.columns[mapped_indices["name"]] == "name"
    source_headers = {column.header: column.index for column in table.source_columns if column.header not in (None, "")}
    assert source_headers["Email"] == mapped_indices["email"]
    assert source_headers["Name"] == mapped_indices["name"]
    score_columns = {table.table.columns[index]: scores for index, scores in table.column_scores.items()}
    assert score_columns["email"] == {"email": 1.0}
    assert score_columns["name"] == {"name": 1.0}

    emitted = list(output_ws.iter_rows(min_row=1, max_row=4, max_col=2, values_only=True))
    assert emitted == [
        ("email", "name"),
        ("a@example.com", None),
        (None, None),
        (None, "Alice"),
    ]


def test_merge_tables_in_sheet_coerces_null_and_string_columns():
    registry = Registry()
    logger = NullLogger()

    def detector(*, row_index, **_):
        if row_index in (1, 4):
            return {RowKind.HEADER.value: 1.0}
        if row_index in (2, 5):
            return {RowKind.DATA.value: 1.0}
        return {}

    registry.register_row_detector(detector, row_kind=RowKind.UNKNOWN.value, priority=0)
    registry.finalize()

    pipeline = Pipeline(
        registry=registry,
        settings=Settings(merge_tables_in_sheet=True, write_diagnostics_columns=False),
        logger=logger,
    )

    source_wb = Workbook()
    source_ws = source_wb.active
    source_ws.title = "Sheet1"
    source_ws.append(["SIN", "Comments"])
    source_ws.append(["123456789", None])
    source_ws.append([])
    source_ws.append(["SIN", "Comments"])
    source_ws.append(["987654321", "NQ"])

    output_wb = Workbook()
    output_wb.remove(output_wb.active)
    output_ws = output_wb.create_sheet(title="Sheet1")

    tables = pipeline.process_sheet(
        sheet=source_ws,
        output_sheet=output_ws,
        state={},
        metadata={"input_file": "apr.xlsx", "sheet_index": 0},
        input_file_name="apr.xlsx",
    )

    assert len(tables) == 1
    assert tables[0].table.columns[0:2] == ["SIN", "Comments"]
    assert tables[0].table.schema["Comments"] == pl.String

    emitted = list(output_ws.iter_rows(min_row=1, max_row=4, max_col=2, values_only=True))
    assert emitted == [
        ("SIN", "Comments"),
        ("123456789", None),
        (None, None),
        ("987654321", "NQ"),
    ]


def test_merge_tables_in_sheet_widens_numeric_types():
    registry = Registry()
    logger = NullLogger()

    def detector(*, row_index, **_):
        if row_index in (1, 4):
            return {RowKind.HEADER.value: 1.0}
        if row_index in (2, 5):
            return {RowKind.DATA.value: 1.0}
        return {}

    registry.register_row_detector(detector, row_kind=RowKind.UNKNOWN.value, priority=0)
    registry.finalize()

    pipeline = Pipeline(
        registry=registry,
        settings=Settings(merge_tables_in_sheet=True, write_diagnostics_columns=False),
        logger=logger,
    )

    source_wb = Workbook()
    source_ws = source_wb.active
    source_ws.title = "Sheet1"
    source_ws.append(["Amount"])
    source_ws.append([10])
    source_ws.append([])
    source_ws.append(["Amount"])
    source_ws.append([2.5])

    output_wb = Workbook()
    output_wb.remove(output_wb.active)
    output_ws = output_wb.create_sheet(title="Sheet1")

    tables = pipeline.process_sheet(
        sheet=source_ws,
        output_sheet=output_ws,
        state={},
        metadata={"input_file": "input.xlsx", "sheet_index": 0},
        input_file_name="input.xlsx",
    )

    assert len(tables) == 1
    assert tables[0].table.schema["Amount"] == pl.Float64
    assert tables[0].table.get_column("Amount").to_list() == [10.0, None, 2.5]


def test_merge_tables_in_sheet_raises_clear_error_for_incompatible_types():
    registry = Registry()
    logger = NullLogger()

    def detector(*, row_index, **_):
        if row_index in (1, 4):
            return {RowKind.HEADER.value: 1.0}
        if row_index in (2, 5):
            return {RowKind.DATA.value: 1.0}
        return {}

    registry.register_row_detector(detector, row_kind=RowKind.UNKNOWN.value, priority=0)
    registry.finalize()

    pipeline = Pipeline(
        registry=registry,
        settings=Settings(merge_tables_in_sheet=True, write_diagnostics_columns=False),
        logger=logger,
    )

    source_wb = Workbook()
    source_ws = source_wb.active
    source_ws.title = "Sheet1"
    source_ws.append(["Mixed"])
    source_ws.append([True])
    source_ws.append([])
    source_ws.append(["Mixed"])
    source_ws.append(["text"])

    output_wb = Workbook()
    output_wb.remove(output_wb.active)
    output_ws = output_wb.create_sheet(title="Sheet1")

    with pytest.raises(PipelineError, match=r"incompatible dtypes for column 'Mixed'"):
        pipeline.process_sheet(
            sheet=source_ws,
            output_sheet=output_ws,
            state={},
            metadata={"input_file": "input.xlsx", "sheet_index": 0},
            input_file_name="input.xlsx",
        )


def test_on_table_written_receives_merged_metadata_for_merged_tables():
    registry = Registry()
    logger = NullLogger()

    registry.register_field(FieldDef(name="email"))
    registry.register_field(FieldDef(name="name"))

    def detector(*, row_index, **_):
        if row_index in (1, 4):
            return {RowKind.HEADER.value: 1.0}
        if row_index in (2, 5):
            return {RowKind.DATA.value: 1.0}
        return {}

    def detect_email(*, column_header_original, **_):
        return {"email": 1.0} if column_header_original.strip().lower() == "email" else {}

    def detect_name(*, column_header_original, **_):
        return {"name": 1.0} if column_header_original.strip().lower() == "name" else {}

    def capture_written(*, write_table, table_result, state, **_):
        state["written_columns"] = list(write_table.columns)
        state["mapped_columns"] = [(column.field_name, column.source_index) for column in table_result.mapped_columns]
        state["source_columns"] = [(column.header, column.index) for column in table_result.source_columns]
        state["source_region"] = table_result.source_region.a1

    registry.register_row_detector(detector, row_kind=RowKind.UNKNOWN.value, priority=0)
    registry.register_column_detector(detect_email, field="email", priority=0)
    registry.register_column_detector(detect_name, field="name", priority=0)
    registry.register_hook(capture_written, hook="on_table_written", priority=0)
    registry.finalize()

    pipeline = Pipeline(
        registry=registry,
        settings=Settings(merge_tables_in_sheet=True, write_diagnostics_columns=False),
        logger=logger,
    )

    source_wb = Workbook()
    source_ws = source_wb.active
    source_ws.title = "Sheet1"
    source_ws.append(["Email"])
    source_ws.append(["a@example.com"])
    source_ws.append([])
    source_ws.append(["Name"])
    source_ws.append(["Alice"])

    output_wb = Workbook()
    output_wb.remove(output_wb.active)
    output_ws = output_wb.create_sheet(title="Sheet1")

    state: dict = {}
    pipeline.process_sheet(
        sheet=source_ws,
        output_sheet=output_ws,
        state=state,
        metadata={"input_file": "input.xlsx", "sheet_index": 0},
        input_file_name="input.xlsx",
    )

    assert state["written_columns"] == ["email", "name"]
    assert [field for field, _ in state["mapped_columns"]] == ["email", "name"]
    source_headers = {header: index for header, index in state["source_columns"] if header not in (None, "")}
    mapped_indices = dict(state["mapped_columns"])
    assert source_headers["Email"] == mapped_indices["email"]
    assert source_headers["Name"] == mapped_indices["name"]
    assert state["source_region"] == "A1:A5"


def test_merge_tables_in_sheet_preserves_integer_supertype_without_float_widening():
    table_a = TableResult(
        sheet_name="Sheet1",
        table=pl.DataFrame({"member_id": pl.Series("member_id", [1], dtype=pl.Int32)}),
        source_region=TableRegion(min_row=1, min_col=1, max_row=2, max_col=1),
        source_columns=[SourceColumn(index=0, header="Member ID", values=[1])],
        table_index=0,
        sheet_index=0,
        row_count=1,
    )
    table_b = TableResult(
        sheet_name="Sheet1",
        table=pl.DataFrame({"member_id": pl.Series("member_id", [2], dtype=pl.Int64)}),
        source_region=TableRegion(min_row=4, min_col=1, max_row=5, max_col=1),
        source_columns=[SourceColumn(index=0, header="Member ID", values=[2])],
        table_index=1,
        sheet_index=0,
        row_count=1,
    )

    merged = _merge_tables_in_sheet([table_a, table_b])[0]

    assert merged.table.schema["member_id"] == pl.Int64
    assert merged.table.get_column("member_id").to_list() == [1, 2]


def test_merge_tables_in_sheet_preserves_duplicate_unmapped_indices():
    table_a = TableResult(
        sheet_name="Sheet1",
        table=pl.DataFrame({"Notes": ["alpha"]}),
        source_region=TableRegion(min_row=1, min_col=1, max_row=2, max_col=1),
        source_columns=[SourceColumn(index=0, header="Notes", values=["alpha"])],
        table_index=0,
        sheet_index=0,
        unmapped_columns=[SourceColumn(index=0, header="Notes", values=["alpha"])],
        duplicate_unmapped_indices={0},
        row_count=1,
    )
    table_b = TableResult(
        sheet_name="Sheet1",
        table=pl.DataFrame({"Notes": ["beta"]}),
        source_region=TableRegion(min_row=4, min_col=1, max_row=5, max_col=1),
        source_columns=[SourceColumn(index=0, header="Notes", values=["beta"])],
        table_index=1,
        sheet_index=0,
        unmapped_columns=[SourceColumn(index=0, header="Notes", values=["beta"])],
        duplicate_unmapped_indices={0},
        row_count=1,
    )

    merged = _merge_tables_in_sheet([table_a, table_b])[0]

    assert merged.duplicate_unmapped_indices == {0}


def test_build_merged_table_result_materializes_each_merged_column_once(monkeypatch: pytest.MonkeyPatch):
    merged_df = pl.DataFrame(
        {
            "email": ["a@example.com", None],
            "name": [None, "Alice"],
        }
    )
    table_a = TableResult(
        sheet_name="Sheet1",
        table=pl.DataFrame({"email": ["a@example.com"]}),
        source_region=TableRegion(min_row=1, min_col=1, max_row=2, max_col=1),
        source_columns=[SourceColumn(index=0, header="Email", values=["a@example.com"])],
        table_index=0,
        sheet_index=0,
        mapped_columns=[
            MappedColumn(
                field_name="email",
                source_index=0,
                header="Email",
                values=["a@example.com"],
                score=0.9,
            )
        ],
        row_count=1,
    )
    table_b = TableResult(
        sheet_name="Sheet1",
        table=pl.DataFrame({"name": ["Alice"]}),
        source_region=TableRegion(min_row=4, min_col=1, max_row=5, max_col=1),
        source_columns=[SourceColumn(index=0, header="Name", values=["Alice"])],
        table_index=1,
        sheet_index=0,
        mapped_columns=[
            MappedColumn(
                field_name="name",
                source_index=0,
                header="Name",
                values=["Alice"],
                score=0.8,
            )
        ],
        row_count=1,
    )

    get_column_calls: dict[str, int] = {}
    original_get_column = pl.DataFrame.get_column

    def counting_get_column(self: pl.DataFrame, name: str) -> pl.Series:
        if self is merged_df:
            get_column_calls[name] = get_column_calls.get(name, 0) + 1
        return original_get_column(self, name)

    monkeypatch.setattr(pl.DataFrame, "get_column", counting_get_column)

    merged = _build_merged_table_result([table_a, table_b], merged_df)

    assert merged.table.columns == ["email", "name"]
    assert get_column_calls == {"email": 1, "name": 1}
