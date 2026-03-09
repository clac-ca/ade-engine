from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from ade_engine.application.engine import Engine
from ade_engine.infrastructure.settings import Settings
from ade_engine.models.run import RunErrorCode, RunRequest, RunStatus


def _write_config_package(root: Path) -> None:
    pkg = root / "ade_config"
    pkg.mkdir(parents=True, exist_ok=True)
    (pkg / "__init__.py").write_text("", encoding="utf-8")

    row_detectors = pkg / "row_detectors"
    row_detectors.mkdir(parents=True, exist_ok=True)
    (row_detectors / "__init__.py").write_text("", encoding="utf-8")
    (row_detectors / "pick_header.py").write_text(
        """
from ade_engine.models import RowKind

def pick_header(*, row_values, **_):
    normalized = {str(v).strip().lower() for v in row_values or [] if v not in (None, "")}
    if {"email", "name"}.issubset(normalized):
        return {"header": 1.0}
    return {}

def register(registry):
    registry.register_row_detector(pick_header, row_kind=RowKind.HEADER.value, priority=10)
""",
        encoding="utf-8",
    )

    columns = pkg / "columns"
    columns.mkdir(parents=True, exist_ok=True)
    (columns / "__init__.py").write_text("", encoding="utf-8")
    (columns / "contact_fields.py").write_text(
        """
import polars as pl
from ade_engine.models import FieldDef

def detect_email(*, column_header_original: str, **_):
    header = (column_header_original or "").strip().lower()
    return {"email": 1.0} if header == "email" else None

def detect_name(*, column_header_original: str, **_):
    header = (column_header_original or "").strip().lower()
    return {"name": 1.0} if header == "name" else None

def normalize_email(*, field_name: str, **_):
    return pl.col(field_name).cast(pl.Utf8).str.to_lowercase()

def validate_email(*, field_name: str, **_):
    v = pl.col(field_name).cast(pl.Utf8)
    return (
        pl.when(v.is_not_null() & ~v.str.contains("@"))
        .then(pl.lit("invalid email"))
        .otherwise(pl.lit(None))
    )

def register(registry):
    registry.register_field(FieldDef(name="email"))
    registry.register_field(FieldDef(name="name"))
    registry.register_column_detector(detect_email, field="email", priority=20)
    registry.register_column_detector(detect_name, field="name", priority=10)
    registry.register_column_transform(normalize_email, field="email", priority=0)
    registry.register_column_validator(validate_email, field="email", priority=0)
""",
        encoding="utf-8",
    )


def _write_xls_workbook(path: Path) -> None:
    xlwt = pytest.importorskip("xlwt")

    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("Sheet1")
    sheet.write(0, 0, "Email")
    sheet.write(0, 1, "Name")
    sheet.write(0, 2, "Notes")
    sheet.write(1, 0, "USER@Example.com")
    sheet.write(1, 1, "Alice")
    sheet.write(1, 2, "note-1")
    sheet.write(2, 0, "bademail")
    sheet.write(2, 1, "Bob")
    sheet.write(2, 2, "note-2")
    workbook.save(str(path))


def test_end_to_end_pipeline_for_xls(tmp_path: Path):
    config_root = tmp_path / "cfg"
    config_root.mkdir()
    _write_config_package(config_root)

    source = tmp_path / "input.xls"
    _write_xls_workbook(source)

    engine = Engine(settings=Settings())
    output_dir = tmp_path / "out"
    logs_dir = tmp_path / "logs"
    req = RunRequest(
        config_package=config_root,
        input_file=source,
        output_dir=output_dir,
        logs_dir=logs_dir,
    )

    result = engine.run(req)
    assert result.status == RunStatus.SUCCEEDED
    output_file = output_dir / f"{source.stem}_normalized.xlsx"
    assert output_file.exists()

    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))[0:3]]
    assert headers == ["email", "name", "Notes"]

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert rows[0][0] == "user@example.com"
    assert rows[0][1] == "Alice"
    assert rows[0][2] == "note-1"
    assert rows[1][0] == "bademail"
    assert rows[1][2] == "note-2"

    wb.close()


def test_corrupt_xls_fails_with_input_error(tmp_path: Path):
    config_root = tmp_path / "cfg"
    config_root.mkdir()
    _write_config_package(config_root)

    source = tmp_path / "broken.xls"
    source.write_text("not an xls file", encoding="utf-8")

    engine = Engine(settings=Settings())
    result = engine.run(
        RunRequest(
            config_package=config_root,
            input_file=source,
            output_dir=tmp_path / "out",
            logs_dir=tmp_path / "logs",
        )
    )

    assert result.status == RunStatus.FAILED
    assert result.error is not None
    assert result.error.code == RunErrorCode.INPUT_ERROR

