from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

from ade_engine.application.engine import Engine
from ade_engine.infrastructure.io import workbook as workbook_io
from ade_engine.infrastructure.settings import Settings
from ade_engine.models.run import RunRequest, RunStatus


def _write_config_package(root: Path) -> None:
    pkg = root / "ade_config"
    pkg.mkdir(parents=True, exist_ok=True)
    (pkg / "__init__.py").write_text("", encoding="utf-8")

    row_detectors = pkg / "row_detectors"
    row_detectors.mkdir(parents=True, exist_ok=True)
    (row_detectors / "__init__.py").write_text("", encoding="utf-8")
    (row_detectors / "pick_header.py").write_text(
        """
from ade_engine.models import RowKind

def pick_header(*, row_values, **_):
    normalized = {str(v).strip().lower() for v in row_values or [] if v not in (None, "")}
    if {"email", "name"}.issubset(normalized):
        return {"header": 1.0}
    return {}

def register(registry):
    registry.register_row_detector(pick_header, row_kind=RowKind.HEADER.value, priority=10)
""",
        encoding="utf-8",
    )

    columns = pkg / "columns"
    columns.mkdir(parents=True, exist_ok=True)
    (columns / "__init__.py").write_text("", encoding="utf-8")
    (columns / "contact_fields.py").write_text(
        """
from ade_engine.models import FieldDef

def detect_email(*, column_header_original: str, **_):
    header = (column_header_original or "").strip().lower()
    return {"email": 1.0} if header == "email" else None

def detect_name(*, column_header_original: str, **_):
    header = (column_header_original or "").strip().lower()
    return {"name": 1.0} if header == "name" else None

def register(registry):
    registry.register_field(FieldDef(name="email"))
    registry.register_field(FieldDef(name="name"))
    registry.register_column_detector(detect_email, field="email", priority=20)
    registry.register_column_detector(detect_name, field="name", priority=10)
""",
        encoding="utf-8",
    )


class _FakeXlsCell:
    def __init__(self, ctype: int, value: Any):
        self.ctype = ctype
        self.value = value


class _FakeXlsSheet:
    def __init__(self, name: str, rows: list[list[_FakeXlsCell]]):
        self.name = name
        self._rows = rows
        self.nrows = len(rows)
        self.ncols = len(rows[0]) if rows else 0

    def cell(self, row_index: int, col_index: int) -> _FakeXlsCell:
        return self._rows[row_index][col_index]


class _FakeXlsBook:
    def __init__(self, sheets: list[_FakeXlsSheet]):
        self._sheets = sheets
        self.datemode = 0

    def sheets(self) -> list[_FakeXlsSheet]:
        return self._sheets


class _FakePdfPage:
    def __init__(self, tables: list[list[list[str | None]]]):
        self._tables = tables

    def extract_tables(self) -> list[list[list[str | None]]]:
        return self._tables


class _FakePdf:
    def __init__(self, pages: list[_FakePdfPage]):
        self.pages = pages

    def __enter__(self) -> "_FakePdf":
        return self

    def __exit__(self, *_: Any) -> bool:
        return False


def test_end_to_end_pipeline_with_xls_input(tmp_path, monkeypatch):
    config_root = tmp_path / "cfg"
    config_root.mkdir()
    _write_config_package(config_root)

    source = tmp_path / "input.xls"
    source.write_bytes(b"fake-xls")

    fake_xlrd = type("FakeXlrd", (), {})()
    fake_xlrd.XL_CELL_EMPTY = 0
    fake_xlrd.XL_CELL_TEXT = 1
    fake_xlrd.XL_CELL_NUMBER = 2
    fake_xlrd.XL_CELL_DATE = 3
    fake_xlrd.XL_CELL_BOOLEAN = 4
    fake_xlrd.XL_CELL_ERROR = 5
    fake_xlrd.XL_CELL_BLANK = 6
    fake_xlrd.xldate_as_datetime = lambda *_: None
    fake_xlrd.open_workbook = lambda **_: _FakeXlsBook(
        sheets=[
            _FakeXlsSheet(
                "Sheet1",
                [
                    [_FakeXlsCell(fake_xlrd.XL_CELL_TEXT, "Email"), _FakeXlsCell(fake_xlrd.XL_CELL_TEXT, "Name")],
                    [_FakeXlsCell(fake_xlrd.XL_CELL_TEXT, "alice@example.com"), _FakeXlsCell(fake_xlrd.XL_CELL_TEXT, "Alice")],
                ],
            )
        ]
    )
    monkeypatch.setattr(workbook_io, "_import_optional_module", lambda name: fake_xlrd if name == "xlrd" else None)

    engine = Engine(settings=Settings())
    output_dir = tmp_path / "out"
    logs_dir = tmp_path / "logs"
    result = engine.run(
        RunRequest(
            config_package=config_root,
            input_file=source,
            output_dir=output_dir,
            logs_dir=logs_dir,
        )
    )

    assert result.status == RunStatus.SUCCEEDED
    output_file = output_dir / "input_normalized.xlsx"
    assert output_file.exists()
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    assert [ws["A1"].value, ws["B1"].value] == ["email", "name"]
    assert [ws["A2"].value, ws["B2"].value] == ["alice@example.com", "Alice"]
    wb.close()


def test_end_to_end_pipeline_with_pdf_input(tmp_path, monkeypatch):
    config_root = tmp_path / "cfg"
    config_root.mkdir()
    _write_config_package(config_root)

    source = tmp_path / "input.pdf"
    source.write_bytes(b"%PDF-1.4 fake")

    fake_pdfplumber = type("FakePdfPlumber", (), {})()
    fake_pdfplumber.open = lambda *_: _FakePdf(
        pages=[_FakePdfPage(tables=[[["Email", "Name"], ["bob@example.com", "Bob"]]])]
    )
    monkeypatch.setattr(
        workbook_io,
        "_import_optional_module",
        lambda name: fake_pdfplumber if name == "pdfplumber" else None,
    )

    engine = Engine(settings=Settings())
    output_dir = tmp_path / "out"
    logs_dir = tmp_path / "logs"
    result = engine.run(
        RunRequest(
            config_package=config_root,
            input_file=source,
            output_dir=output_dir,
            logs_dir=logs_dir,
        )
    )

    assert result.status == RunStatus.SUCCEEDED
    output_file = output_dir / "input_normalized.xlsx"
    assert output_file.exists()
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    assert [ws["A1"].value, ws["B1"].value] == ["email", "name"]
    assert [ws["A2"].value, ws["B2"].value] == ["bob@example.com", "Bob"]
    wb.close()
