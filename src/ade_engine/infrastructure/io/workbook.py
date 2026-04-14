"""Workbook IO helpers for :class:`~ade_engine.application.engine.Engine`."""

from __future__ import annotations

import csv
import math
import re
from io import BytesIO
from contextlib import contextmanager, suppress
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl
from openpyxl import Workbook

from ade_engine.models.errors import InputError

_XLS_METADATA_ATTR = "_ade_xls_metadata"
_XLS_ORIGINAL_NAME_ATTR = "_ade_xls_original_name"


@dataclass(frozen=True)
class _XlsWorkbookMetadata:
    is_converted_xls: bool = True


def load_source_workbook(path: Path) -> Workbook:
    """Load source data from CSV/XLS/XLSX into a workbook."""

    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _load_csv_workbook(path)
    if suffix == ".xls":
        return _load_xls_workbook(path)
    return _load_openpyxl_workbook(path)


def _load_csv_workbook(path: Path) -> Workbook:
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise InputError("Failed to initialize worksheet for CSV input")
    ws.title = path.stem
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.reader(handle):
            ws.append(row)
    return wb


def _load_openpyxl_workbook(path: Path) -> Workbook:
    try:
        return openpyxl.load_workbook(filename=path, read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - openpyxl owns error types
        repaired_workbook = _try_load_workbook_with_repaired_styles(path, exc)
        if repaired_workbook is not None:
            return repaired_workbook
        raise InputError(f"Failed to open workbook '{path}': {exc}") from exc


def _try_load_workbook_with_repaired_styles(path: Path, exc: Exception) -> Workbook | None:
    if not _is_invalid_font_family_stylesheet_error(exc):
        return None

    repaired_bytes = _repair_invalid_font_families_in_archive(path)
    if repaired_bytes is None:
        return None

    try:
        return openpyxl.load_workbook(filename=BytesIO(repaired_bytes), read_only=True, data_only=True)
    except Exception:
        return None


def _is_invalid_font_family_stylesheet_error(exc: Exception) -> bool:
    if "could not read stylesheet" not in str(exc):
        return False

    current: BaseException | None = exc
    seen: set[int] = set()
    while current is not None and id(current) not in seen:
        seen.add(id(current))
        if "Max value is 14" in str(current):
            return True
        current = current.__cause__ or current.__context__
    return False


def _repair_invalid_font_families_in_archive(path: Path) -> bytes | None:
    with ZipFile(path) as archive:
        try:
            styles_xml = archive.read("xl/styles.xml")
        except KeyError:
            return None

        repaired_styles_xml, changed = _remove_invalid_font_family_elements(styles_xml)
        if not changed:
            return None

        buffer = BytesIO()
        with ZipFile(buffer, "w", compression=ZIP_DEFLATED) as repaired_archive:
            for member in archive.infolist():
                data = repaired_styles_xml if member.filename == "xl/styles.xml" else archive.read(member.filename)
                repaired_archive.writestr(member, data)
        return buffer.getvalue()


def _remove_invalid_font_family_elements(styles_xml: bytes) -> tuple[bytes, bool]:
    pattern = re.compile(rb"<family\b[^>]*\bval=\"(\d+)\"[^>]*/>")
    changed = False

    def _replace(match: re.Match[bytes]) -> bytes:
        nonlocal changed
        family_value = int(match.group(1))
        if family_value <= 14:
            return match.group(0)
        changed = True
        return b""

    return pattern.sub(_replace, styles_xml), changed


def _load_xls_workbook(path: Path) -> Workbook:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover - dependency should exist in normal installs
        raise InputError("Failed to open .xls workbook: xlrd is not installed") from exc

    try:
        book = xlrd.open_workbook(
            filename=str(path),
            formatting_info=True,
            on_demand=False,
            ragged_rows=True,
        )
    except Exception as exc:  # pragma: no cover - xlrd owns error types
        raise InputError(f"Failed to open .xls workbook '{path}': {exc}") from exc

    try:
        return _convert_xlrd_book_to_openpyxl(book)
    except InputError:
        raise
    except Exception as exc:  # pragma: no cover - defensive
        raise InputError(f"Failed to convert .xls workbook '{path}': {exc}") from exc


def _convert_xlrd_book_to_openpyxl(book: Any) -> Workbook:
    workbook = Workbook()
    if workbook.worksheets:
        workbook.remove(workbook.worksheets[0])

    visible_sheet_index: int | None = None
    selected_sheet_index: int | None = None
    active_sheet_index: int | None = None
    for sheet_index in range(book.nsheets):
        sheet = book.sheet_by_index(sheet_index)
        ws = workbook.create_sheet(title=sheet.name)
        setattr(ws, _XLS_ORIGINAL_NAME_ATTR, str(sheet.name))

        visibility = _get_xls_sheet_visibility(book, sheet, sheet_index)
        ws.sheet_state = _map_sheet_visibility(visibility)
        is_visible = ws.sheet_state == "visible"
        is_active = bool(getattr(sheet, "sheet_visible", False))
        is_selected = bool(getattr(sheet, "sheet_selected", False))
        if visible_sheet_index is None and is_visible:
            visible_sheet_index = sheet_index
        if active_sheet_index is None and is_visible and is_active:
            active_sheet_index = sheet_index
        if selected_sheet_index is None and is_visible and is_selected:
            selected_sheet_index = sheet_index

        for row_index in range(sheet.nrows):
            row_values = [_convert_xls_cell_value(book, cell) for cell in _iter_xls_row_cells(sheet, row_index)]
            ws.append(row_values)

        for merged_range in getattr(sheet, "merged_cells", ()):
            rlo, rhi, clo, chi = merged_range
            if rhi - rlo <= 1 and chi - clo <= 1:
                continue
            ws.merge_cells(
                start_row=rlo + 1,
                end_row=rhi,
                start_column=clo + 1,
                end_column=chi,
            )

    if not workbook.worksheets:
        raise InputError("Workbook contains no worksheets")

    resolved_active_index = active_sheet_index
    if resolved_active_index is None:
        resolved_active_index = selected_sheet_index
    if resolved_active_index is None:
        resolved_active_index = visible_sheet_index
    if resolved_active_index is None:
        resolved_active_index = 0

    workbook.active = resolved_active_index
    _set_xls_workbook_metadata(workbook)
    return workbook


def _map_sheet_visibility(visibility: int) -> str:
    if visibility == 1:
        return "hidden"
    if visibility == 2:
        return "veryHidden"
    return "visible"


def _iter_xls_row_cells(sheet: Any, row_index: int) -> list[Any]:
    row_len = getattr(sheet, "row_len", None)
    row_types = getattr(sheet, "row_types", None)
    row_values = getattr(sheet, "row_values", None)

    if callable(row_len) and callable(row_types) and callable(row_values):
        row_cell_types = row_types(row_index, 0, row_len(row_index))
        row_cell_values = row_values(row_index, 0, row_len(row_index))
        return [
            SimpleNamespace(ctype=cell_type, value=cell_value)
            for cell_type, cell_value in zip(row_cell_types, row_cell_values, strict=True)
        ]

    return [sheet.cell(row_index, col_index) for col_index in range(sheet.ncols)]


def _get_xls_sheet_visibility(book: Any, sheet: Any, sheet_index: int) -> int:
    visibility = getattr(sheet, "visibility", None)
    if isinstance(visibility, int):
        return visibility

    visibility_by_sheet = getattr(book, "_sheet_visibility", None)
    if visibility_by_sheet is None:
        visibility_by_sheet = getattr(book, "sheet_visibility", None)
    if isinstance(visibility_by_sheet, (list, tuple)) and sheet_index < len(visibility_by_sheet):
        raw_visibility = visibility_by_sheet[sheet_index]
        if isinstance(raw_visibility, int):
            return raw_visibility
    return 0


def _set_xls_workbook_metadata(workbook: Workbook) -> None:
    setattr(
        workbook,
        _XLS_METADATA_ATTR,
        _XlsWorkbookMetadata(),
    )


def _get_xls_workbook_metadata(workbook: Workbook) -> _XlsWorkbookMetadata | None:
    metadata = getattr(workbook, _XLS_METADATA_ATTR, None)
    if isinstance(metadata, _XlsWorkbookMetadata):
        return metadata
    return None


def _convert_xls_cell_value(book: Any, cell: Any) -> Any:
    try:
        import xlrd
        from xlrd.biffh import error_text_from_code
    except ImportError as exc:  # pragma: no cover - dependency should exist in normal installs
        raise InputError("Failed to convert .xls workbook: xlrd is not installed") from exc

    cell_type = cell.ctype
    value = cell.value

    if cell_type in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return None
    if cell_type == xlrd.XL_CELL_TEXT:
        return str(value)
    if cell_type == xlrd.XL_CELL_BOOLEAN:
        return bool(value)
    if cell_type == xlrd.XL_CELL_ERROR:
        return error_text_from_code.get(value)
    if cell_type == xlrd.XL_CELL_DATE:
        converted = xlrd.xldate_as_datetime(value, book.datemode)
        if isinstance(converted, datetime):
            return converted
        return converted
    if cell_type == xlrd.XL_CELL_NUMBER:
        numeric = float(value)
        if math.isfinite(numeric) and numeric.is_integer():
            return int(numeric)
        return numeric
    return value


@contextmanager
def open_source_workbook(path: Path):
    """Context manager for safely opening source workbooks."""

    workbook = load_source_workbook(path)
    try:
        yield workbook
    finally:
        with suppress(Exception):
            workbook.close()


def create_output_workbook() -> Workbook:
    """Create a clean output workbook with no default sheet."""

    workbook = Workbook()
    if workbook.worksheets:
        workbook.remove(workbook.worksheets[0])
    return workbook


def resolve_sheet_names(
    workbook: Workbook,
    requested: list[str] | None,
    *,
    active_only: bool = False,
) -> list[str]:
    """Determine which sheets to process, preserving source order."""

    xls_metadata = _get_xls_workbook_metadata(workbook)
    if xls_metadata is not None:
        return _resolve_xls_sheet_names(workbook, requested, active_only=active_only)

    visible = [ws.title for ws in workbook.worksheets if getattr(ws, "sheet_state", "visible") == "visible"]
    if active_only:
        if not visible:
            raise InputError("No visible worksheets available")
        active = workbook.active
        active_name = getattr(active, "title", None)
        if not active_name:
            raise InputError("Active worksheet is not available")
        if active_name not in visible:
            raise InputError(f"Active worksheet is hidden: {active_name}")
        return [active_name]
    if not requested:
        return visible

    cleaned = [name.strip() for name in requested if isinstance(name, str) and name.strip()]
    unique_requested = list(dict.fromkeys(cleaned))  # preserve order, drop duplicates

    missing = [name for name in unique_requested if name not in visible]
    if missing:
        raise InputError(f"Worksheet(s) not found: {', '.join(missing)}")

    order_index = {name: idx for idx, name in enumerate(visible)}
    return sorted(unique_requested, key=lambda n: order_index[n])


def _resolve_xls_sheet_names(
    workbook: Workbook,
    requested: list[str] | None,
    *,
    active_only: bool,
) -> list[str]:
    visible_sheets = [ws for ws in workbook.worksheets if getattr(ws, "sheet_state", "visible") == "visible"]
    visible_titles = [ws.title for ws in visible_sheets]
    if active_only:
        if not visible_sheets:
            raise InputError("No visible worksheets available")
        active = workbook.active
        active_name = getattr(active, "title", None)
        if not active_name:
            raise InputError("Active worksheet is not available")
        if active_name not in visible_titles:
            raise InputError(f"Active worksheet is hidden: {active_name}")
        return [active_name]

    if not requested:
        return visible_titles

    cleaned = [name.strip() for name in requested if isinstance(name, str) and name.strip()]
    unique_requested = list(dict.fromkeys(cleaned))

    visible_by_requested_name: dict[str, str] = {}
    for ws in visible_sheets:
        visible_by_requested_name[ws.title] = ws.title
        visible_by_requested_name[str(getattr(ws, _XLS_ORIGINAL_NAME_ATTR, ws.title))] = ws.title

    missing = [name for name in unique_requested if name not in visible_by_requested_name]
    if missing:
        raise InputError(f"Worksheet(s) not found: {', '.join(missing)}")

    order_index = {ws.title: idx for idx, ws in enumerate(visible_sheets)}
    resolved = [visible_by_requested_name[name] for name in unique_requested]
    return sorted(resolved, key=lambda n: order_index[n])


__all__ = [
    "create_output_workbook",
    "load_source_workbook",
    "open_source_workbook",
    "resolve_sheet_names",
]
